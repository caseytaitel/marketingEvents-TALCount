#!/usr/bin/env python3
"""Build TAL Accounts @ Marketing Events from a Marketing Events Registry."""

from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from hubspot import (
    load_token,
    make_session,
    normalize_event_token,
    primary_companies_for_contacts,
    search_contacts_by_events_attended,
    tal_flags_for_companies,
    tokens_for_contact,
)

REGISTRY_SHEET = "Registry"
CHANNEL_CATEGORY = "Channel Partner Events"

SHEET_ALL = "TAL Accounts @Events"
SHEET_CHANNEL = "TAL Accounts @Channel Events"

HEADERS = [
    "Category",
    "Sub-category",
    "Events Attended",
    "# Contact Attendees",
    "# Unique Accounts",
    "# Accounts on TAL",
    "% Accounts on TAL",
]

COLUMN_WIDTHS = {
    "A": 18.75,
    "B": 24.38,
    "C": 55.5,
    "D": 17.13,
    "E": 16.13,
    "F": 22.5,
    "G": 16.63,
}

FONT_ARIAL = Font(name="Arial", size=10)
FONT_ARIAL_BOLD = Font(name="Arial", size=10, bold=True)
FONT_TOTAL = Font(name="Arial", size=12, bold=True)
FILL_HEADER = PatternFill("solid", fgColor="C9DAF8")
FILL_TOTAL = PatternFill("solid", fgColor="FFF2CC")
FILL_TOTAL_PCT = PatternFill("solid", fgColor="D9D2E9")
ALIGN_BOTTOM = Alignment(vertical="bottom")
ALIGN_CENTER = Alignment(horizontal="center", vertical="bottom")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="bottom")
ROW_HEIGHT = 15.75
PCT_COLOR_SCALE = ColorScaleRule(
    start_type="min",
    start_color="E67C73",
    mid_type="percentile",
    mid_value=50,
    mid_color="FFFFFF",
    end_type="max",
    end_color="57BB8A",
)


@dataclass(frozen=True)
class RegistryRow:
    kind: str  # "event" | "blank"
    category: str
    subcategory: str
    appendage: str


@dataclass(frozen=True)
class Metrics:
    attendees: int
    unique_accounts: int
    accounts_on_tal: int


def main(argv: list[str] | None = None) -> int:
    args = _parse_args(argv)
    token = load_token()
    registry_path = args.registry.expanduser().resolve()
    out_path = _resolve_out_path(args.out, registry_path)

    rows = read_registry(registry_path)
    event_rows = [row for row in rows if row.kind == "event"]
    appendages = [row.appendage for row in event_rows]

    session = make_session(token)
    contacts = search_contacts_by_events_attended(session, appendages)
    contacts_by_token = _index_contacts_by_token(contacts)

    unmatched = _unmatched_appendages(event_rows, contacts_by_token)
    if unmatched:
        print("No HubSpot contacts matched these Events Attended Appendage values:", file=sys.stderr)
        for value in unmatched:
            print(f"  - {value}", file=sys.stderr)
        print("Wrote nothing.", file=sys.stderr)
        return 1

    all_union = _union_contact_ids(event_rows, contacts_by_token)
    channel_rows = [
        row for row in event_rows if row.category.strip() == CHANNEL_CATEGORY
    ]
    channel_union = _union_contact_ids(channel_rows, contacts_by_token)

    primary = primary_companies_for_contacts(session, list(all_union))
    company_ids = list(dict.fromkeys(primary.values()))
    tal = tal_flags_for_companies(session, company_ids)

    per_event = {
        normalize_event_token(row.appendage): _metrics_for_contacts(
            contacts_by_token.get(normalize_event_token(row.appendage), set()),
            primary,
            tal,
        )
        for row in event_rows
    }
    all_total = _metrics_for_contacts(all_union, primary, tal)
    channel_total = _metrics_for_contacts(channel_union, primary, tal)

    _print_console(event_rows, per_event, all_total, channel_total)

    wb = Workbook()
    _set_workbook_defaults(wb)
    ws_all = wb.active
    ws_all.title = SHEET_ALL
    ws_channel = wb.create_sheet(SHEET_CHANNEL)
    _write_all_events_sheet(ws_all, rows, per_event, all_total)
    _write_channel_sheet(ws_channel, channel_rows, per_event, channel_total)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Wrote {out_path}")
    return 0


def _parse_args(argv: list[str] | None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Write TAL Accounts @ Marketing Events from a Marketing Events Registry."
    )
    parser.add_argument(
        "registry",
        nargs="?",
        type=Path,
        default=Path("INPUT.xlsx"),
        help="Path to the Marketing Events Registry .xlsx (default: INPUT.xlsx)",
    )
    parser.add_argument(
        "--out",
        type=Path,
        default=None,
        help="Output .xlsx path (default: outputs/YYYY-MM-DD/tal_accounts_marketing_events.xlsx)",
    )
    return parser.parse_args(argv)


def _resolve_out_path(out_arg: Path | None, registry_path: Path) -> Path:
    if out_arg is None:
        out_path = Path("outputs") / date.today().isoformat() / "tal_accounts_marketing_events.xlsx"
    else:
        out_path = out_arg.expanduser()
        if out_path.suffix.lower() != ".xlsx":
            out_path = out_path.with_suffix(".xlsx")
    resolved = out_path.resolve()
    if resolved == registry_path:
        raise SystemExit("Output path matches the registry; refusing to overwrite.")
    return resolved


def read_registry(path: Path) -> list[RegistryRow]:
    if not path.is_file():
        raise SystemExit(f"Registry file not found: {path}")
    wb = load_workbook(path, data_only=True)
    try:
        if REGISTRY_SHEET not in wb.sheetnames:
            raise SystemExit(
                f"Workbook has no sheet named {REGISTRY_SHEET!r}. Found: {wb.sheetnames}"
            )
        ws = wb[REGISTRY_SHEET]
        _require_headers(ws)

        raw_rows: list[tuple[str, str, str]] = []
        for excel_row in ws.iter_rows(min_row=2, max_col=7, values_only=True):
            category = _display(excel_row[0] if excel_row else None)
            subcategory = _display(excel_row[1] if excel_row and len(excel_row) > 1 else None)
            appendage = _display(excel_row[6] if excel_row and len(excel_row) > 6 else None)
            raw_rows.append((category, subcategory, appendage))
    finally:
        wb.close()

    last_event = -1
    for i, (_category, _subcategory, appendage) in enumerate(raw_rows):
        if appendage.strip():
            last_event = i
    if last_event < 0:
        return []

    included: list[RegistryRow] = []
    for category, subcategory, appendage in raw_rows[: last_event + 1]:
        if appendage.strip():
            included.append(
                RegistryRow(
                    kind="event",
                    category=category,
                    subcategory=subcategory,
                    appendage=appendage,
                )
            )
        elif not category.strip() and not subcategory.strip():
            included.append(
                RegistryRow(kind="blank", category="", subcategory="", appendage="")
            )
    return included


def _require_headers(ws: Worksheet) -> None:
    found = {
        "A": _display(ws["A1"].value).strip(),
        "B": _display(ws["B1"].value).strip(),
        "G": _display(ws["G1"].value).strip(),
    }
    expected = {
        "A": "Category",
        "B": "Sub-category",
        "G": "Events Attended Appendage",
    }
    mismatches = [
        f"{col} (expected {expected[col]!r}, found {found[col]!r})"
        for col in expected
        if found[col] != expected[col]
    ]
    if mismatches:
        raise SystemExit(
            "Registry headers do not match expected columns: " + "; ".join(mismatches)
        )


def _display(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _index_contacts_by_token(contacts: list[dict]) -> dict[str, set[str]]:
    index: dict[str, set[str]] = {}
    for contact in contacts:
        contact_id = str(contact["id"])
        props = contact.get("properties") or {}
        for token in tokens_for_contact(props.get("events_attended")):
            index.setdefault(token, set()).add(contact_id)
    return index


def _unmatched_appendages(
    event_rows: list[RegistryRow], contacts_by_token: dict[str, set[str]]
) -> list[str]:
    unmatched: list[str] = []
    seen: set[str] = set()
    for row in event_rows:
        token = normalize_event_token(row.appendage)
        if contacts_by_token.get(token):
            continue
        if row.appendage in seen:
            continue
        seen.add(row.appendage)
        unmatched.append(row.appendage)
    return unmatched


def _union_contact_ids(
    event_rows: list[RegistryRow], contacts_by_token: dict[str, set[str]]
) -> set[str]:
    union: set[str] = set()
    for row in event_rows:
        union |= contacts_by_token.get(normalize_event_token(row.appendage), set())
    return union


def _metrics_for_contacts(
    contact_ids: set[str],
    primary: dict[str, str],
    tal: dict[str, bool],
) -> Metrics:
    companies = {primary[cid] for cid in contact_ids if cid in primary}
    tal_count = sum(1 for company_id in companies if tal.get(company_id))
    return Metrics(
        attendees=len(contact_ids),
        unique_accounts=len(companies),
        accounts_on_tal=tal_count,
    )


def _print_console(
    event_rows: list[RegistryRow],
    per_event: dict[str, Metrics],
    all_total: Metrics,
    channel_total: Metrics,
) -> None:
    for row in event_rows:
        metrics = per_event[normalize_event_token(row.appendage)]
        print(
            f"{row.appendage}\tD={metrics.attendees}\t"
            f"E={metrics.unique_accounts}\tF={metrics.accounts_on_tal}"
        )
    print(
        f"TOTAL (all events)\tD={all_total.attendees}\t"
        f"E={all_total.unique_accounts}\tF={all_total.accounts_on_tal}"
    )
    print(
        f"TOTAL (Channel Partner Events)\tD={channel_total.attendees}\t"
        f"E={channel_total.unique_accounts}\tF={channel_total.accounts_on_tal}"
    )


def _write_all_events_sheet(
    ws: Worksheet,
    rows: list[RegistryRow],
    per_event: dict[str, Metrics],
    total: Metrics,
) -> None:
    _apply_column_widths(ws)
    _write_header(ws)
    excel_row = 2
    for row in rows:
        if row.kind == "blank":
            _style_identity_row(ws, excel_row)
            excel_row += 1
            continue
        metrics = per_event[normalize_event_token(row.appendage)]
        _write_event_row(ws, excel_row, row, metrics)
        excel_row += 1
    _write_total_row(ws, excel_row, total)
    _finish_sheet(ws)


def _write_channel_sheet(
    ws: Worksheet,
    channel_rows: list[RegistryRow],
    per_event: dict[str, Metrics],
    total: Metrics,
) -> None:
    _apply_column_widths(ws)
    _write_header(ws)
    excel_row = 2
    for row in channel_rows:
        metrics = per_event[normalize_event_token(row.appendage)]
        _write_event_row(ws, excel_row, row, metrics)
        excel_row += 1
    _style_identity_row(ws, excel_row)
    excel_row += 1
    _write_total_row(ws, excel_row, total)
    _finish_sheet(ws)


def _set_workbook_defaults(wb: Workbook) -> None:
    default = wb._fonts[0]
    default.name = "Arial"
    default.sz = 10.0
    default.family = 2
    default.scheme = "minor"


def _finish_sheet(ws: Worksheet) -> None:
    ws.freeze_panes = "A2"
    ws.sheet_format.defaultRowHeight = ROW_HEIGHT
    ws.sheet_format.customHeight = True
    last = ws.max_row or 1
    for row in range(1, last + 1):
        ws.row_dimensions[row].height = ROW_HEIGHT
    if last >= 2:
        ws.conditional_formatting.add(f"G2:G{last}", PCT_COLOR_SCALE)


def _apply_column_widths(ws: Worksheet) -> None:
    for letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[letter].width = width


def _write_header(ws: Worksheet) -> None:
    for col, title in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = FONT_ARIAL_BOLD
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_BOTTOM


def _style_identity_row(ws: Worksheet, excel_row: int) -> None:
    for col in range(1, 8):
        cell = ws.cell(row=excel_row, column=col)
        cell.font = FONT_ARIAL
        cell.alignment = ALIGN_BOTTOM


def _write_event_row(
    ws: Worksheet, excel_row: int, row: RegistryRow, metrics: Metrics
) -> None:
    values = {
        1: row.category,
        2: row.subcategory,
        3: row.appendage,
        4: metrics.attendees,
        5: metrics.unique_accounts,
        6: metrics.accounts_on_tal,
    }
    for col, value in values.items():
        cell = ws.cell(row=excel_row, column=col, value=value)
        cell.font = FONT_ARIAL
        cell.alignment = ALIGN_CENTER if col >= 4 else ALIGN_BOTTOM
        if col >= 4:
            cell.number_format = "0"
    pct = ws.cell(row=excel_row, column=7, value=f"=F{excel_row}/E{excel_row}")
    pct.font = FONT_ARIAL
    pct.alignment = ALIGN_CENTER
    pct.number_format = "0%"


def _write_total_row(ws: Worksheet, excel_row: int, metrics: Metrics) -> None:
    for col in range(1, 3):
        cell = ws.cell(row=excel_row, column=col)
        cell.font = FONT_ARIAL
        cell.alignment = ALIGN_BOTTOM
    label = ws.cell(row=excel_row, column=3, value="TOTAL")
    label.font = FONT_TOTAL
    label.fill = FILL_TOTAL
    label.alignment = ALIGN_RIGHT
    for col, value in ((4, metrics.attendees), (5, metrics.unique_accounts), (6, metrics.accounts_on_tal)):
        cell = ws.cell(row=excel_row, column=col, value=value)
        cell.font = FONT_TOTAL
        cell.fill = FILL_TOTAL
        cell.alignment = ALIGN_CENTER
        cell.number_format = "0"
    pct = ws.cell(row=excel_row, column=7, value=f"=F{excel_row}/E{excel_row}")
    pct.font = FONT_TOTAL
    pct.fill = FILL_TOTAL_PCT
    pct.alignment = ALIGN_CENTER
    pct.number_format = "0.00%"


if __name__ == "__main__":
    sys.exit(main())
